import sys
sys.stdout.reconfigure(encoding='utf-8')

import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = r"C:/Users/Admin/Desktop/beldoc/private/uploads/baremes/1781293006272-barema-new-01042026.xlsx"

# --- Pass 1: values (data_only) ---
wb = load_workbook(PATH, data_only=True)
print("SHEETNAMES:", wb.sheetnames)
ws = wb["Bonus"]
print(f"DIMENSIONS: {ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")

print("\n=== NON-EMPTY CELLS (data_only=True) ===")
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if v is None:
            continue
        t = type(v).__name__
        if isinstance(v, str) and v.strip() == "":
            t += " (blank-string)"
        rv = repr(v)
        print(f"{cell.coordinate}\t[{t}]\tnumfmt={cell.number_format!r}\t{rv}")

# --- Pass 2: formulas + merged cells (no data_only) ---
wb2 = load_workbook(PATH, data_only=False)
ws2 = wb2["Bonus"]
print("\n=== MERGED RANGES ===")
for mr in sorted(ws2.merged_cells.ranges, key=lambda m: (m.min_row, m.min_col)):
    print(mr)

print("\n=== FORMULA CELLS (raw, data_only=False) ===")
for row in ws2.iter_rows():
    for cell in row:
        v = cell.value
        if v is None:
            continue
        if isinstance(v, str) and v.startswith("="):
            print(f"{cell.coordinate}\tFORMULA\t{v!r}")

# --- Numeric amount candidates ---
print("\n=== NUMERIC CELLS (data_only) ===")
count = 0
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            count += 1
            print(f"{cell.coordinate}\t{v}\tnumfmt={cell.number_format!r}")
        elif isinstance(v, datetime.datetime) or isinstance(v, datetime.date):
            print(f"{cell.coordinate}\tDATE\t{v}")
print(f"TOTAL numeric cells: {count}")

# --- Column widths / hidden? ---
print("\n=== COLUMN DIMENSIONS (hidden?) ===")
for key, dim in ws2.column_dimensions.items():
    print(f"col {key}: width={dim.width} hidden={dim.hidden}")
print("\n=== ROW DIMENSIONS (hidden?) ===")
for key, dim in ws2.row_dimensions.items():
    if dim.hidden:
        print(f"row {key}: HIDDEN")
